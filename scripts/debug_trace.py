"""Batch eval — runs every question from All Questions.xlsx through the RAG graph
and writes results to C:\\Users\\HC162DJ\\Downloads\\report.xlsx.

Columns written:
  SL No | Question | Answer | Citation FileName | Page Numbers | Page Content
  | Citation Block | Groundedness Score | Relevance Score
  | Answer Completeness Score | Safety Score | Ground Truth Score

Run:
    python -m scripts.debug_trace
"""
import asyncio
import logging
import os
import re
import sys
from pathlib import Path

os.environ["PYTHONIOENCODING"] = "utf-8"

# Suppress all noisy logs — only show script progress
logging.basicConfig(level=logging.ERROR, format="%(name)s | %(message)s")

import openpyxl
from langchain_core.messages import HumanMessage
from agents.rag.graph import build_rag_graph

INPUT_XLSX  = r"C:\Users\HC162DJ\Downloads\All Questions.xlsx"
OUTPUT_XLSX = r"C:\Users\HC162DJ\Downloads\report.xlsx"

HEADERS = [
    "SL No", "Question", "Answer",
    "Citation FileName", "Page Numbers", "Page Content",
    "Citation Block",
    "Groundedness Score", "Relevance Score",
    "Answer Completeness Score", "Safety Score", "Ground Truth Score",
]

_FALLBACK = (
    "I couldn't find a specific answer to your question in "
    "the available knowledge base. Please try rephrasing your "
    "query or provide more details."
)

_STOPWORDS = frozenset({
    "the","a","an","is","are","was","were","be","been","being","have","has",
    "had","do","does","did","will","would","could","should","may","might",
    "shall","can","need","must","in","on","at","to","for","with","by","from",
    "of","about","and","or","but","not","no","nor","so","yet","both","this",
    "that","these","those","it","its","they","them","he","she","his","her",
    "we","our","you","your",
})


def _tokenize(text: str) -> set[str]:
    words = re.findall(r"\b[a-z0-9]+(?:[-/][a-z0-9]+)*\b", text.lower())
    return {w for w in words if w not in _STOPWORDS and len(w) > 2}


def _groundedness(answer: str, events: list) -> float:
    if not events or not answer:
        return 0.0
    atoks = _tokenize(answer)
    if not atoks:
        return 0.0
    dtoks: set[str] = set()
    for ev in events:
        dtoks |= _tokenize(ev.get("content", ""))
    if not dtoks:
        return 0.0
    return round(len(atoks & dtoks) / len(atoks), 3)


def _completeness(answer: str) -> float:
    if not answer or answer.strip() == _FALLBACK.strip():
        return 0.0
    return 1.0


def _load_questions() -> list[tuple]:
    wb = openpyxl.load_workbook(INPUT_XLSX, read_only=True)
    ws = wb.active
    rows = [(r[0], str(r[1]).strip()) for r in ws.iter_rows(min_row=2, values_only=True) if r[1]]
    wb.close()
    return rows


def _init_workbook():
    """Load existing report (for resume) or create fresh one."""
    if Path(OUTPUT_XLSX).exists():
        wb = openpyxl.load_workbook(OUTPUT_XLSX)
        ws = wb.active
        done = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
        return wb, ws, done
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(HEADERS)
    wb.save(OUTPUT_XLSX)
    return wb, ws, set()


async def _run_query(graph, question: str) -> dict:
    state = {
        "messages": [HumanMessage(content=question)],
        "user_input": question,
        "user_id": "eval@ey.com",
        "chat_id": None,
        "chat_session_id": None,
        "message_id": None,
        "input_type": "ask",
        "is_free_form": True,
        "function": [],
        "sub_function": [],
        "source_url": [],
        "start_date": "",
        "end_date": "",
        "preferred_language": "English",
        "content_type": "document",
        "requires_function_selection": False,
        "channel_type": 0,
    }
    config = {"configurable": {"thread_id": f"eval-{abs(hash(question))}"}}

    answer, events = "", []
    async for chunk in graph.astream(state, config=config, stream_mode="updates"):
        for node_name, node_output in chunk.items():
            if node_name in ("generate", "synthesize"):
                ai = node_output.get("ai_content", "")
                if ai:
                    answer = ai
                evts = node_output.get("events", [])
                if evts:
                    events = evts

    citation_block = ""
    m = re.search(r"\n\nCitations:\n([\s\S]+)$", answer)
    if m:
        citation_block = m.group(0).strip()

    return {"answer": answer, "events": events, "citation_block": citation_block}


async def main():
    questions = _load_questions()
    print(f"Total questions: {len(questions)}")

    wb, ws, done = _init_workbook()
    print(f"Already done: {len(done)}, remaining: {len(questions) - len(done)}")

    graph = build_rag_graph(checkpointer=None, memory_store=None)

    for i, (sl, question) in enumerate(questions):
        if sl in done:
            continue

        print(f"[{i+1}/{len(questions)}] SL {sl}: {question[:70]}", flush=True)

        try:
            res = await _run_query(graph, question)
        except Exception as exc:
            print(f"  ERROR: {exc}", flush=True)
            res = {"answer": f"ERROR: {exc}", "events": [], "citation_block": ""}

        answer         = res["answer"]
        events         = res["events"]
        citation_block = res["citation_block"]

        file_names = "; ".join(dict.fromkeys(
            ev.get("file_name", "") for ev in events if ev.get("file_name")
        ))
        page_nums = "; ".join(dict.fromkeys(
            str(ev.get("page_number", ""))
            for ev in events if ev.get("page_number") not in (None, "")
        ))
        page_content = " | ".join(
            ev.get("content", "")[:200].replace("\n", " ") for ev in events[:3]
        )

        ws.append([
            sl,
            question,
            answer,
            file_names,
            page_nums,
            page_content,
            citation_block,
            _groundedness(answer, events),   # Groundedness Score
            "N/A",                           # Relevance Score (reranker not in merged events)
            _completeness(answer),           # Answer Completeness Score
            "N/A",                           # Safety Score
            "N/A",                           # Ground Truth Score
        ])
        wb.save(OUTPUT_XLSX)

    print(f"\nDone. Report saved to: {OUTPUT_XLSX}")


if __name__ == "__main__":
    asyncio.run(main())

