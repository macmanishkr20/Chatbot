"""
Excel → UserExpenses ETL loader.

Source: the "MENA Expense Report" Concur export. Each line item on a
report becomes one row in ``UserExpenses``. The DDL lives in
``sql/create_agent_tables.sql`` and is owned by DBAs — this loader only
inserts data.

Idempotency: re-running the same source file produces the same DB
state. The loader queries existing rows, computes the same dedupe
projection (employee + report + transaction-date + transaction-amount
+ transaction-currency + vendor + business-purpose), and skips any
candidate row that already matches an existing row.

Usage:
    python -m tools.data_ingestion.expense_loader <path.xlsx> --triggered-by user@ey.com

Header matching is case-insensitive, whitespace-tolerant, and accepts
the most common Concur export variants. Override via ``column_map``.
"""
from __future__ import annotations

import argparse
import asyncio
import logging
import math
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Mapping

from models.data_models import ImportRunResult
from services.data_db import DataDB

logger = logging.getLogger(__name__)


# Canonical key → set of accepted Concur header variants
# (lower-cased, whitespace-collapsed). Add new variants here as they appear.
_DEFAULT_COLUMN_MAP: dict[str, tuple[str, ...]] = {
    "country_name":             ("country name", "country"),
    "country_code":             ("country code",),
    "company_code":             ("company code",),
    "company_code_description": ("company code description", "company description"),
    "cost_center_id":           ("cost center id", "cost center code", "cost centre id"),
    "cost_center":              ("cost center", "cost centre"),
    "employee_id":              ("employee id", "employee_id", "emp id", "empid", "gpn"),
    "employee_name":            ("employee name", "name"),
    "home_address":             ("home address",),
    "employee_rank":            ("employee rank", "rank", "band"),
    "report_id":                ("report id",),
    "report_key":               ("report key",),
    "report_name":              ("report name",),
    "policy":                   ("policy",),
    "approval_status":          ("approval status",),
    "approved_by":              ("approved by", "approver"),
    "payment_status":           ("payment status",),
    "trip_start_date":          ("trip start date",),
    "trip_end_date":            ("trip end date",),
    "original_submission_datetime":   (
        "original submission date/time", "original submission date time",
        "original submission datetime", "original submission date",
    ),
    "last_submitted_datetime":        (
        "last submitted date/time", "last submitted date time",
        "last submitted datetime", "last submitted date",
    ),
    "approval_status_change_datetime":(
        "approval status change date/time", "approval status change date time",
        "approval status change datetime", "approval status change date",
    ),
    "payment_status_change_date":     ("payment status change date",),
    "transaction_date":         ("transaction date", "txn date"),
    "expense_type":             ("expense type",),
    "expense_sub_type1":        ("expense sub-type 1", "expense sub type 1", "expense subtype 1"),
    "expense_sub_type2":        ("expense sub-type 2", "expense sub type 2", "expense subtype 2"),
    "origin":                   ("origin",),
    "destination":              ("destination",),
    "from_date":                ("from date",),
    "to_date":                  ("to date",),
    "business_purpose":         ("business purpose",),
    "original_reimbursement_amount": ("original reimbursement amount",),
    "reimbursement_amount":     ("reimbursement amount",),
    "reimbursement_currency":   ("reimbursement currency",),
    "transaction_amount":       ("transaction amount",),
    "transaction_currency":     ("transaction currency",),
    "work_location_country":    ("work location country",),
    "work_location_region":     ("work location region", "work location state", "work location province"),
    "work_location_city":       ("work location city",),
    "country_of_purchase":      ("country of purchase",),
    "region_of_purchase":       ("region of purchase", "state of purchase"),
    "city_of_purchase":         ("city of purchase",),
    "vendor":                   ("vendor", "merchant", "supplier"),
    "receipt_status":           ("receipt status",),
    "gl_account":               ("gl account", "g/l account"),
    "engagement_name":          ("engagement name",),
    "engagement_code":          ("engagement code",),
    "engagement_percentage":    ("engagement percentage", "engagement %"),
    "transaction_type":         ("transaction type", "txn type"),
    "number_of_attendees":      ("number of attendees", "no of attendees", "attendees"),
    "trip_over_3_months":       ("trip over 3 months",),
}


def _norm_header(s: Any) -> str:
    return " ".join(str(s or "").strip().split()).lower()


def _normalise_columns(headers: list[str]) -> dict[str, str]:
    """Map our canonical names → actual sheet header (preserving case)."""
    lookup = {_norm_header(h): h for h in headers if h is not None}
    resolved: dict[str, str] = {}
    for canonical, candidates in _DEFAULT_COLUMN_MAP.items():
        for c in candidates:
            if c in lookup:
                resolved[canonical] = lookup[c]
                break
    return resolved


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return Decimal(str(value))
    if isinstance(value, Decimal):
        return value
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d-%b-%Y", "%d-%b-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    s = str(value).strip()
    if not s:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M", "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y",
        "%d-%b-%Y %H:%M:%S", "%d-%b-%Y", "%d-%b-%y",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _dedupe_projection(record: Mapping[str, Any]) -> tuple:
    """Stable dedupe key projected from a parsed record. Mirrors the
    SQL projection in ``_existing_dedupe_projections`` exactly."""
    td = record.get("transaction_date")
    td_iso = (
        td.isoformat() if isinstance(td, (date, datetime)) else ""
    )
    amt = record.get("transaction_amount")
    amount_cents = int((amt or Decimal("0")) * 100)
    return (
        str(record.get("employee_id") or "").strip().lower(),
        str(record.get("report_id") or "").strip().lower(),
        td_iso,
        amount_cents,
        str(record.get("transaction_currency") or "").strip().upper(),
        str(record.get("vendor") or "").strip().lower(),
        str(record.get("business_purpose") or "").strip().lower(),
    )


def _read_workbook(path: Path) -> tuple[list[str], list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required for Excel ingestion. "
            "Install it with `pip install openpyxl`.",
        ) from exc

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return headers, [list(r) for r in rows[1:]]


async def _existing_dedupe_projections(db: DataDB) -> set[tuple]:
    """Read the dedupe projection for every UserExpenses row.

    For typical row counts this scans the table once per import — which
    is acceptable: imports are admin-triggered and infrequent. Add a
    materialised hash column later if the table grows large enough that
    this scan becomes expensive.
    """
    rows = await db.fetchall(
        """
        SELECT EmployeeId, ReportId, TransactionDate,
               TransactionAmount, TransactionCurrency,
               Vendor, BusinessPurpose
        FROM UserExpenses
        """,
    )
    out: set[tuple] = set()
    for r in rows:
        td = r.get("TransactionDate")
        td_iso = td.isoformat() if isinstance(td, (date, datetime)) else ""
        amt = r.get("TransactionAmount")
        amount_cents = int((Decimal(str(amt)) if amt is not None else Decimal("0")) * 100)
        out.add((
            (r.get("EmployeeId") or "").strip().lower(),
            (r.get("ReportId") or "").strip().lower(),
            td_iso,
            amount_cents,
            (r.get("TransactionCurrency") or "").strip().upper(),
            (r.get("Vendor") or "").strip().lower(),
            (r.get("BusinessPurpose") or "").strip().lower(),
        ))
    return out


async def import_expense_workbook(
    workbook_path: str | Path,
    *,
    triggered_by: str | None = None,
    column_map: dict[str, str] | None = None,
) -> ImportRunResult:
    """Load an Excel workbook into UserExpenses. Idempotent.

    Args:
        workbook_path: path to the .xlsx file.
        triggered_by:  stamped on the audit row.
        column_map:    override the canonical → sheet-header mapping.
    """
    path = Path(workbook_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(path)

    db = DataDB()
    await db.ensure()  # creates audit tables only

    run_id_row = await db.fetchone(
        """
        INSERT INTO ExpenseImportRuns (SourceFile, TriggeredBy)
        OUTPUT INSERTED.RunId
        VALUES (?, ?)
        """,
        [str(path.name), triggered_by],
    )
    run_id = (run_id_row or {}).get("RunId")

    headers, raw_rows = _read_workbook(path)
    resolved_cols = column_map or _normalise_columns(headers)

    # Minimum viable columns to anchor each row + power dedupe.
    required = {"employee_id", "transaction_date", "transaction_amount"}
    if not required.issubset(resolved_cols):
        await db.execute(
            """
            UPDATE ExpenseImportRuns
            SET Status='failed', ErrorMessage=?, FinishedAt=SYSUTCDATETIME()
            WHERE RunId=?
            """,
            [f"Missing required columns. Resolved: {sorted(resolved_cols)}", run_id],
        )
        raise ValueError(
            "Workbook is missing required columns "
            "(employee_id / transaction_date / transaction_amount). "
            f"Found columns: {headers}"
        )

    header_idx = {h: i for i, h in enumerate(headers)}

    def col(row: list[Any], canonical: str) -> Any:
        sheet_header = resolved_cols.get(canonical)
        if sheet_header is None:
            return None
        idx = header_idx.get(sheet_header)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    # ── Pass 1: parse rows ──
    parsed: list[dict[str, Any]] = []
    rows_skipped = 0
    for raw in raw_rows:
        emp = col(raw, "employee_id")
        td = _parse_datetime(col(raw, "transaction_date"))
        if not emp or td is None:
            rows_skipped += 1
            continue

        record: dict[str, Any] = {
            # Identity
            "employee_id": str(emp).strip(),
            "employee_name": col(raw, "employee_name"),
            "home_address": col(raw, "home_address"),
            "employee_rank": col(raw, "employee_rank"),
            # Org
            "country_name": col(raw, "country_name"),
            "country_code": col(raw, "country_code"),
            "company_code": col(raw, "company_code"),
            "company_code_description": col(raw, "company_code_description"),
            "cost_center_id": col(raw, "cost_center_id"),
            "cost_center": col(raw, "cost_center"),
            # Report
            "report_id": col(raw, "report_id"),
            "report_key": _parse_int(col(raw, "report_key")),
            "report_name": col(raw, "report_name"),
            "policy": col(raw, "policy"),
            "approval_status": col(raw, "approval_status"),
            "approved_by": col(raw, "approved_by"),
            "payment_status": col(raw, "payment_status"),
            # Dates
            "trip_start_date": _parse_date(col(raw, "trip_start_date")),
            "trip_end_date": _parse_date(col(raw, "trip_end_date")),
            "original_submission_datetime": _parse_datetime(col(raw, "original_submission_datetime")),
            "last_submitted_datetime": _parse_datetime(col(raw, "last_submitted_datetime")),
            "approval_status_change_datetime": _parse_datetime(col(raw, "approval_status_change_datetime")),
            "payment_status_change_date": _parse_datetime(col(raw, "payment_status_change_date")),
            "transaction_date": td,
            # Categorisation
            "expense_type": col(raw, "expense_type"),
            "expense_sub_type1": col(raw, "expense_sub_type1"),
            "expense_sub_type2": col(raw, "expense_sub_type2"),
            # Trip
            "origin": col(raw, "origin"),
            "destination": col(raw, "destination"),
            "from_date": _parse_date(col(raw, "from_date")),
            "to_date": _parse_date(col(raw, "to_date")),
            "business_purpose": col(raw, "business_purpose"),
            # Money
            "original_reimbursement_amount": _parse_decimal(col(raw, "original_reimbursement_amount")),
            "reimbursement_amount": _parse_decimal(col(raw, "reimbursement_amount")),
            "reimbursement_currency": (str(col(raw, "reimbursement_currency") or "").strip().upper() or None),
            "transaction_amount": _parse_decimal(col(raw, "transaction_amount")),
            "transaction_currency": (str(col(raw, "transaction_currency") or "").strip().upper() or None),
            # Locations
            "work_location_country": col(raw, "work_location_country"),
            "work_location_region": col(raw, "work_location_region"),
            "work_location_city": col(raw, "work_location_city"),
            "country_of_purchase": col(raw, "country_of_purchase"),
            "region_of_purchase": col(raw, "region_of_purchase"),
            "city_of_purchase": col(raw, "city_of_purchase"),
            # Misc
            "vendor": col(raw, "vendor"),
            "receipt_status": col(raw, "receipt_status"),
            "gl_account": col(raw, "gl_account"),
            "engagement_name": col(raw, "engagement_name"),
            "engagement_code": col(raw, "engagement_code"),
            "engagement_percentage": _parse_decimal(col(raw, "engagement_percentage")),
            "transaction_type": col(raw, "transaction_type"),
            "number_of_attendees": _parse_int(col(raw, "number_of_attendees")),
            "trip_over_3_months": col(raw, "trip_over_3_months"),
        }
        parsed.append(record)

    # ── Pass 2: filter out duplicates already in the table ──
    existing = await _existing_dedupe_projections(db)
    fresh = [r for r in parsed if _dedupe_projection(r) not in existing]
    rows_skipped += len(parsed) - len(fresh)

    # ── Pass 3: insert ──
    insert_sql = """
        INSERT INTO UserExpenses (
            CountryName, CountryCode, CompanyCode, CompanyCodeDescription,
            CostCenterId, CostCenter, EmployeeId, EmployeeName, HomeAddress,
            EmployeeRank, ReportId, ReportKey, ReportName, Policy,
            ApprovalStatus, ApprovedBy, PaymentStatus,
            TripStartDate, TripEndDate,
            OriginalSubmissionDateTime, LastSubmittedDateTime,
            ApprovalStatusChangeDateTime, PaymentStatusChangeDate,
            TransactionDate, ExpenseType, ExpenseSubType1, ExpenseSubType2,
            Origin, Destination, FromDate, ToDate, BusinessPurpose,
            OriginalReimbursementAmount, ReimbursementAmount, ReimbursementCurrency,
            TransactionAmount, TransactionCurrency,
            WorkLocationCountry, WorkLocationRegion, WorkLocationCity,
            CountryOfPurchase, RegionOfPurchase, CityOfPurchase,
            Vendor, ReceiptStatus, GLAccount,
            EngagementName, EngagementCode, EngagementPercentage,
            TransactionType, NumberOfAttendees, TripOver3Months
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """
    rows_for_insert: list[tuple[Any, ...]] = []
    for r in fresh:
        rows_for_insert.append((
            r["country_name"], r["country_code"], r["company_code"], r["company_code_description"],
            r["cost_center_id"], r["cost_center"], r["employee_id"], r["employee_name"], r["home_address"],
            r["employee_rank"], r["report_id"], r["report_key"], r["report_name"], r["policy"],
            r["approval_status"], r["approved_by"], r["payment_status"],
            r["trip_start_date"], r["trip_end_date"],
            r["original_submission_datetime"], r["last_submitted_datetime"],
            r["approval_status_change_datetime"], r["payment_status_change_date"],
            r["transaction_date"], r["expense_type"], r["expense_sub_type1"], r["expense_sub_type2"],
            r["origin"], r["destination"], r["from_date"], r["to_date"], r["business_purpose"],
            r["original_reimbursement_amount"], r["reimbursement_amount"], r["reimbursement_currency"],
            r["transaction_amount"], r["transaction_currency"],
            r["work_location_country"], r["work_location_region"], r["work_location_city"],
            r["country_of_purchase"], r["region_of_purchase"], r["city_of_purchase"],
            r["vendor"], r["receipt_status"], r["gl_account"],
            r["engagement_name"], r["engagement_code"], r["engagement_percentage"],
            r["transaction_type"], r["number_of_attendees"], r["trip_over_3_months"],
        ))

    inserted = await db.execute_many(insert_sql, rows_for_insert)

    await db.execute(
        """
        UPDATE ExpenseImportRuns
        SET RowsRead=?, RowsInserted=?, RowsSkipped=?, Status='ok',
            FinishedAt=SYSUTCDATETIME()
        WHERE RunId=?
        """,
        [len(raw_rows), inserted, rows_skipped, run_id],
    )

    return ImportRunResult(
        run_id=run_id,
        source_file=path.name,
        rows_read=len(raw_rows),
        rows_inserted=inserted,
        rows_skipped=rows_skipped,
        status="ok",
    )


def _cli() -> None:  # pragma: no cover
    parser = argparse.ArgumentParser(description="Import a MENA Expense Report Excel into UserExpenses.")
    parser.add_argument("workbook", help="Path to the .xlsx file.")
    parser.add_argument("--triggered-by", default=None, help="User identifier for audit row.")
    args = parser.parse_args()
    result = asyncio.run(
        import_expense_workbook(args.workbook, triggered_by=args.triggered_by),
    )
    print(result.model_dump_json(indent=2))


if __name__ == "__main__":  # pragma: no cover
    _cli()
