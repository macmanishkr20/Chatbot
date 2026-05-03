"""
Excel → UserScoreboard ETL loader.

Source: the "MENA Scorecard data_Template_v3" workbook, Template sheet.
Wide format: each row is one (employee, period) record with all KPI
columns inline. The DDL lives in ``sql/create_agent_tables.sql``.

Idempotency: re-running the same workbook is a no-op. The dedupe
projection is (EmployeeId | Period | ReportDate); rows already present
are skipped.

Usage:
    python -m tools.data_ingestion.scoreboard_loader <path.xlsx> --triggered-by user@ey.com
"""
from __future__ import annotations

import argparse
import asyncio
import logging
import math
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from models.data_models import ImportRunResult
from services.data_db import DataDB

logger = logging.getLogger(__name__)


# Canonical key → accepted header variants (lower-cased, whitespace-collapsed)
_DEFAULT_COLUMN_MAP: dict[str, tuple[str, ...]] = {
    "gui":               ("gui",),
    "gpn":               ("gpn",),
    "employee_name":     ("employee name", "name"),
    "employee_id":       ("employee id", "employeeid", "emp id"),
    "country":           ("country",),
    "sl":                ("sl", "service line"),
    "ssl":               ("ssl", "sub service line", "sub-service line"),
    "current_rank":      ("current rank", "rank"),
    "role":              ("role",),
    "additional_role":   ("additional role",),
    # KPIs
    "gter":                  ("gter",),
    "gter_plan":             ("gter plan", "gterplan"),
    "gter_plan_achieved_pct":("gter plan achieved %", "gter plan achieved pct", "gterplanachievedpct"),
    "global_margin":         ("global margin", "globalmargin"),
    "global_margin_pct":     ("global margin %", "global margin pct", "globalmarginpct"),
    "global_sales":          ("global sales", "globalsales"),
    "weighted_pipeline":     ("weighted pipeline", "weightedpipeline"),
    "ter":                   ("ter",),
    "ansr":                  ("ansr",),
    "ansr_gter_ratio":       ("ansr / gter ratio", "ansr/gter ratio", "ansrgterratio", "ansr gter ratio"),
    "eng_margin":            ("eng margin", "engmargin", "engagement margin"),
    "eng_margin_pct":        ("eng margin %", "eng margin pct", "engmarginpct"),
    "fytd_backlog_ter":      ("fytd backlog ter", "fytdbacklogter"),
    "total_backlog_ter":     ("total backlog ter", "totalbacklogter"),
    "utilization_pct":       ("utilization %", "utilization pct", "utilizationpct", "utilisation %"),
    "billing":               ("billing",),
    "collection":            ("collection",),
    "ar":                    ("ar", "accounts receivable"),
    "ar_reserve":            ("ar reserve", "arreserve"),
    "total_nui":             ("total nui", "totalnui"),
    "aged_nui_above_180_days": ("aged nui above 180 days", "agednuiabove180days", ">180 days nui"),
    "aged_nui_above_365_days": ("aged nui above 365 days", "agednuiabove365days", ">365 days nui"),
    "revenue_days":          ("revenue days", "revenuedays"),
    # Period
    "period":                ("period",),
    "report_date":           ("report date", "reportdate", "snapshot date", "as of date"),
}


def _norm_header(s: Any) -> str:
    return " ".join(str(s or "").strip().split()).lower()


def _normalise_columns(headers: list[str]) -> dict[str, str]:
    lookup = {_norm_header(h): h for h in headers if h is not None}
    resolved: dict[str, str] = {}
    for canonical, candidates in _DEFAULT_COLUMN_MAP.items():
        for c in candidates:
            if c in lookup:
                resolved[canonical] = lookup[c]
                break
    return resolved


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return Decimal(str(value))
    if isinstance(value, Decimal):
        return value
    s = str(value).strip().replace(",", "")
    if s.endswith("%"):
        s = s[:-1].strip()
    if not s:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d-%b-%Y", "%d-%b-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _read_workbook(path: Path) -> tuple[list[str], list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required for Excel ingestion. "
            "Install it with `pip install openpyxl`.",
        ) from exc

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return headers, [list(r) for r in rows[1:]]


def _dedupe_projection(record: dict[str, Any]) -> tuple:
    rd = record.get("report_date")
    rd_iso = rd.isoformat() if isinstance(rd, (date, datetime)) else ""
    return (
        str(record.get("employee_id") or "").strip().lower(),
        str(record.get("period") or "").strip().upper(),
        rd_iso,
    )


async def _existing_dedupe_projections(db: DataDB) -> set[tuple]:
    rows = await db.fetchall(
        "SELECT EmployeeId, Period, ReportDate FROM UserScoreboard",
    )
    out: set[tuple] = set()
    for r in rows:
        rd = r.get("ReportDate")
        rd_iso = rd.isoformat() if isinstance(rd, (date, datetime)) else ""
        out.add((
            (r.get("EmployeeId") or "").strip().lower(),
            (r.get("Period") or "").strip().upper(),
            rd_iso,
        ))
    return out


async def import_scoreboard_workbook(
    workbook_path: str | Path,
    *,
    triggered_by: str | None = None,
    column_map: dict[str, str] | None = None,
) -> ImportRunResult:
    """Load a MENA Scorecard workbook into UserScoreboard. Idempotent."""
    path = Path(workbook_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(path)

    db = DataDB()
    await db.ensure()

    run_id_row = await db.fetchone(
        """
        INSERT INTO ScoreboardImportRuns (SourceFile, TriggeredBy)
        OUTPUT INSERTED.RunId
        VALUES (?, ?)
        """,
        [str(path.name), triggered_by],
    )
    run_id = (run_id_row or {}).get("RunId")

    headers, raw_rows = _read_workbook(path)
    resolved_cols = column_map or _normalise_columns(headers)

    required = {"employee_id", "period"}
    if not required.issubset(resolved_cols):
        await db.execute(
            """
            UPDATE ScoreboardImportRuns
            SET Status='failed', ErrorMessage=?, FinishedAt=SYSUTCDATETIME()
            WHERE RunId=?
            """,
            [f"Missing required columns. Resolved: {sorted(resolved_cols)}", run_id],
        )
        raise ValueError(
            "Workbook is missing required columns "
            "(employee_id / period). "
            f"Found columns: {headers}"
        )

    header_idx = {h: i for i, h in enumerate(headers)}

    def col(row: list[Any], canonical: str) -> Any:
        sheet_header = resolved_cols.get(canonical)
        if sheet_header is None:
            return None
        idx = header_idx.get(sheet_header)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    parsed: list[dict[str, Any]] = []
    rows_skipped = 0
    for raw in raw_rows:
        emp = col(raw, "employee_id") or col(raw, "gui")
        period = col(raw, "period")
        if not emp or not period:
            rows_skipped += 1
            continue

        record = {
            "gui": col(raw, "gui"),
            "gpn": col(raw, "gpn"),
            "employee_name": col(raw, "employee_name"),
            "employee_id": str(emp).strip(),
            "country": col(raw, "country"),
            "sl": col(raw, "sl"),
            "ssl": col(raw, "ssl"),
            "current_rank": col(raw, "current_rank"),
            "role": col(raw, "role"),
            "additional_role": col(raw, "additional_role"),
            "gter": _parse_decimal(col(raw, "gter")),
            "gter_plan": _parse_decimal(col(raw, "gter_plan")),
            "gter_plan_achieved_pct": _parse_decimal(col(raw, "gter_plan_achieved_pct")),
            "global_margin": _parse_decimal(col(raw, "global_margin")),
            "global_margin_pct": _parse_decimal(col(raw, "global_margin_pct")),
            "global_sales": _parse_decimal(col(raw, "global_sales")),
            "weighted_pipeline": _parse_decimal(col(raw, "weighted_pipeline")),
            "ter": _parse_decimal(col(raw, "ter")),
            "ansr": _parse_decimal(col(raw, "ansr")),
            "ansr_gter_ratio": _parse_decimal(col(raw, "ansr_gter_ratio")),
            "eng_margin": _parse_decimal(col(raw, "eng_margin")),
            "eng_margin_pct": _parse_decimal(col(raw, "eng_margin_pct")),
            "fytd_backlog_ter": _parse_decimal(col(raw, "fytd_backlog_ter")),
            "total_backlog_ter": _parse_decimal(col(raw, "total_backlog_ter")),
            "utilization_pct": _parse_decimal(col(raw, "utilization_pct")),
            "billing": _parse_decimal(col(raw, "billing")),
            "collection": _parse_decimal(col(raw, "collection")),
            "ar": _parse_decimal(col(raw, "ar")),
            "ar_reserve": _parse_decimal(col(raw, "ar_reserve")),
            "total_nui": _parse_decimal(col(raw, "total_nui")),
            "aged_nui_above_180_days": _parse_decimal(col(raw, "aged_nui_above_180_days")),
            "aged_nui_above_365_days": _parse_decimal(col(raw, "aged_nui_above_365_days")),
            "revenue_days": _parse_decimal(col(raw, "revenue_days")),
            "period": str(period).strip(),
            "report_date": _parse_date(col(raw, "report_date")),
        }
        parsed.append(record)

    existing = await _existing_dedupe_projections(db)
    fresh = [r for r in parsed if _dedupe_projection(r) not in existing]
    rows_skipped += len(parsed) - len(fresh)

    insert_sql = """
        INSERT INTO UserScoreboard (
            GUI, GPN, EmployeeName, EmployeeId,
            Country, SL, SSL, CurrentRank, Role, AdditionalRole,
            GTER, GTERPlan, GTERPlanAchievedPct,
            GlobalMargin, GlobalMarginPct, GlobalSales, WeightedPipeline,
            TER, ANSR, ANSRGTERRatio,
            EngMargin, EngMarginPct,
            FYTDBacklogTER, TotalBacklogTER,
            UtilizationPct, Billing, Collection, AR, ARReserve,
            TotalNUI, AgedNUIAbove180Days, AgedNUIAbove365Days,
            RevenueDays, Period, ReportDate
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """
    rows_for_insert: list[tuple[Any, ...]] = []
    for r in fresh:
        rows_for_insert.append((
            r["gui"], r["gpn"], r["employee_name"], r["employee_id"],
            r["country"], r["sl"], r["ssl"], r["current_rank"], r["role"], r["additional_role"],
            r["gter"], r["gter_plan"], r["gter_plan_achieved_pct"],
            r["global_margin"], r["global_margin_pct"], r["global_sales"], r["weighted_pipeline"],
            r["ter"], r["ansr"], r["ansr_gter_ratio"],
            r["eng_margin"], r["eng_margin_pct"],
            r["fytd_backlog_ter"], r["total_backlog_ter"],
            r["utilization_pct"], r["billing"], r["collection"], r["ar"], r["ar_reserve"],
            r["total_nui"], r["aged_nui_above_180_days"], r["aged_nui_above_365_days"],
            r["revenue_days"], r["period"], r["report_date"],
        ))

    inserted = await db.execute_many(insert_sql, rows_for_insert)

    await db.execute(
        """
        UPDATE ScoreboardImportRuns
        SET RowsRead=?, RowsInserted=?, RowsSkipped=?, Status='ok',
            FinishedAt=SYSUTCDATETIME()
        WHERE RunId=?
        """,
        [len(raw_rows), inserted, rows_skipped, run_id],
    )

    return ImportRunResult(
        run_id=run_id,
        source_file=path.name,
        rows_read=len(raw_rows),
        rows_inserted=inserted,
        rows_skipped=rows_skipped,
        status="ok",
    )


def _cli() -> None:  # pragma: no cover
    parser = argparse.ArgumentParser(description="Import a MENA Scorecard Excel into UserScoreboard.")
    parser.add_argument("workbook", help="Path to the .xlsx file.")
    parser.add_argument("--triggered-by", default=None)
    args = parser.parse_args()
    result = asyncio.run(import_scoreboard_workbook(args.workbook, triggered_by=args.triggered_by))
    print(result.model_dump_json(indent=2))


if __name__ == "__main__":  # pragma: no cover
    _cli()
