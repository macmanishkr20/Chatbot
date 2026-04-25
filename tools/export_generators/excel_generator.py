"""Excel generator — uses openpyxl. Template optional."""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="2E2E38")  # EY charcoal
_HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Arial", size=10, color="2E2E38")
_TITLE_FONT = Font(name="Arial", size=16, bold=True, color="2E2E38")


def build_excel(
    output_path: Path,
    plan: dict[str, Any],
    template_path: Path | None = None,
) -> Path:
    """Build an .xlsx file from the LLM plan.

    Plan shape:
      {
        "title": str,
        "sheets": [
          {"name": str, "headers": [str], "rows": [[cell, ...]], "summary": str?}
        ]
      }
    """
    if template_path and template_path.exists():
        wb = load_workbook(template_path)
        # Drop default empty sheets if any
        for s in list(wb.sheetnames):
            if wb[s].max_row <= 1 and wb[s].max_column <= 1:
                del wb[s]
    else:
        wb = Workbook()
        # Remove the default sheet — we'll add our own
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    title = plan.get("title", "Report")
    sheets = plan.get("sheets") or [{"name": "Data", "headers": [], "rows": []}]

    for sheet_def in sheets:
        name = (sheet_def.get("name") or "Sheet")[:31]
        ws = wb.create_sheet(title=name)

        ws["A1"] = title
        ws["A1"].font = _TITLE_FONT
        ws["A1"].alignment = Alignment(vertical="center")

        headers = sheet_def.get("headers") or []
        rows = sheet_def.get("rows") or []
        start_row = 3

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=str(header))
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, row in enumerate(rows, start=start_row + 1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = _BODY_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col_idx in range(1, max(len(headers), 1) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

        summary = sheet_def.get("summary")
        if summary:
            sum_row = start_row + len(rows) + 2
            ws.cell(row=sum_row, column=1, value="Summary:").font = Font(bold=True)
            ws.cell(row=sum_row + 1, column=1, value=summary).alignment = Alignment(wrap_text=True)

    if not wb.sheetnames:
        wb.create_sheet("Sheet1")

    wb.save(output_path)
    return output_path
