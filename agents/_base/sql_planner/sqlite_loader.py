"""
xlsx → SQLite :memory: loader used by stub data sources for Expense / Scorecard.

The stub backends load the Excel templates into an in-memory SQLite table so
the SAME compiled SQL works in development and production. SQLite's ``TOP``
emulation differs from SQL Server, so we rewrite ``SELECT TOP (N) …`` to
``SELECT … LIMIT N`` at execute time.

Excel headers are mapped to canonical SQL column names via an explicit
mapping dict supplied by the caller. Anything not in the mapping is ignored.
"""
from __future__ import annotations

import logging
import re
import sqlite3
import threading
from datetime import date, datetime
from pathlib import Path
from typing import Any, Sequence

import openpyxl

from agents._base.sql_planner.data_source import DataSourceError

logger = logging.getLogger(__name__)


# Cache: one in-memory DB per (xlsx_path, table_name) tuple
_DB_CACHE: dict[tuple[str, str], sqlite3.Connection] = {}
_DB_LOCK = threading.Lock()


def _coerce_cell(v: Any) -> Any:
    """Map openpyxl cell values to SQLite-friendly types."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def load_sqlite_from_xlsx(
    xlsx_path: str | Path,
    *,
    sheet_name: str,
    table_name: str,
    column_map: dict[str, str],
) -> sqlite3.Connection:
    """Load an Excel sheet into an in-memory SQLite table and return the conn.

    Args:
        xlsx_path:    Path to the .xlsx file.
        sheet_name:   Worksheet name to load.
        table_name:   Target SQLite table name (must be a plain identifier).
        column_map:   {excel_header: sql_column}. Headers not in the map
                      are skipped. SQL columns absent from the workbook
                      are created as NULL.

    Cached by (path, table_name) — repeated calls return the same in-memory
    connection so the agent doesn't reload on every request.
    """
    key = (str(Path(xlsx_path).resolve()), table_name)
    with _DB_LOCK:
        cached = _DB_CACHE.get(key)
        if cached is not None:
            return cached

        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        except FileNotFoundError as e:
            raise DataSourceError(
                code="DATASET_MISSING",
                detail=f"Could not open {xlsx_path}: {e}",
                retriable=False,
            ) from e

        if sheet_name not in wb.sheetnames:
            raise DataSourceError(
                code="SHEET_MISSING",
                detail=f"Sheet {sheet_name!r} not in {xlsx_path} (found {wb.sheetnames}).",
                retriable=False,
            )
        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

        # Build column projection: list of (excel_idx, sql_col)
        cols: list[tuple[int, str]] = []
        for idx, h in enumerate(headers):
            sql_name = column_map.get(h)
            if sql_name:
                cols.append((idx, sql_name))

        if not cols:
            raise DataSourceError(
                code="EMPTY_COLUMN_MAP",
                detail=f"No mapped columns found in {sheet_name!r}; check column_map keys against headers.",
                retriable=False,
            )

        conn = sqlite3.connect(":memory:", check_same_thread=False)
        conn.row_factory = sqlite3.Row

        # All columns typed as TEXT in SQLite — the compiler binds parameters
        # via ``?`` and SQLite coerces. Numerics still compare correctly
        # because we ALWAYS use ``CAST(x AS REAL)`` in aggregates? No — we
        # let SQLite type affinity handle this; numeric strings sort/compare
        # numerically when paired with a numeric parameter.
        # Use NUMERIC affinity so SUM/AVG / comparisons just work.
        col_defs = ", ".join(f'"{c}" NUMERIC' for _, c in cols)
        conn.executescript(f'DROP TABLE IF EXISTS "{table_name}"; CREATE TABLE "{table_name}" ({col_defs});')

        placeholders = ", ".join("?" for _ in cols)
        col_names_quoted = ", ".join(f'"{c}"' for _, c in cols)
        insert_sql = f'INSERT INTO "{table_name}" ({col_names_quoted}) VALUES ({placeholders})'

        rows_inserted = 0
        batch: list[Sequence[Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            batch.append(tuple(_coerce_cell(row[idx]) if idx < len(row) else None for idx, _ in cols))
            if len(batch) >= 500:
                conn.executemany(insert_sql, batch)
                rows_inserted += len(batch)
                batch.clear()
        if batch:
            conn.executemany(insert_sql, batch)
            rows_inserted += len(batch)
        conn.commit()
        wb.close()

        logger.info(
            "SQLite stub: loaded %d rows into %s from %s::%s",
            rows_inserted, table_name, xlsx_path, sheet_name,
        )
        _DB_CACHE[key] = conn
        return conn


# ── Per-execute SQL adapter: rewrite SQL-Server-isms to SQLite ──

_TOP_RE = re.compile(r"^\s*SELECT\s+TOP\s*\(\s*(\d+)\s*\)\s+", re.IGNORECASE)


def sqlserver_to_sqlite(sql: str) -> str:
    """Strip SQL Server flavour to make compiler output runnable on SQLite.

    The compiler emits ``SELECT TOP (N) …`` (T-SQL). SQLite doesn't support
    TOP, so we move the limit to a trailing ``LIMIT N``. Identifiers are
    already plain ASCII (whitelisted), and we never bracket-quote, so no
    other rewrites are needed.
    """
    m = _TOP_RE.match(sql)
    if not m:
        return sql
    limit_n = m.group(1)
    rest = sql[m.end():]
    return f"SELECT {rest} LIMIT {limit_n}"
